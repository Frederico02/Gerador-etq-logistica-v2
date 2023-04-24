#/*            Resete de Spool local SAP - v2      */
#/*		Gupo Multi 20/04/2023            */
#/*	Frederico Almeida - Analista de Suporte PL */

#Bibliotecas Usadas
import socket
import openpyxl
import tkinter as tk
import tkinter.simpledialog as sd
from tkinter import Canvas
from tkinter import messagebox


#Conexão Com a impressora de REDE
def impressora(codigo_zpl):
    # Endereço IP da impressora
    HOST = '192.168.152.52'
    # Porta padrão de comunicação com a impressora
    PORT = 9100

    # Conecta-se à impressora
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.connect((HOST, PORT))

    # Envia o código ZPL para a impressora
    s.sendall(codigo_zpl.encode())

    # Fecha a conexão com a impressora
    s.close()

#Layout Etiqueta
def codigo_impressora(concatena, formatado):
    codigo_zpl = 'CT~~CD,~CC^~CT~\n' \
                 '^XA~TA000~JSN^LT0^MNW^MTT^PON^PMN^LH0,0^JMA^PR4,4~SD20^JUS^LRN^CI0^XZ\n' \
                 '^XA \n' \
                 '^MMT \n' \
                 '^PW559 \n' \
                 '^LL0320 \n' \
                 '^LS0\n' \
                 '^BY2,3,75^FT128,241^BCN,,N,N\n' \
                 '^FD>;' + concatena + '^FS\n' \
                 '^FT63,115^A0N,56,60^FH\^FD' + formatado + '^FS\n' \
                 '^PQ1,0,1,Y^XZ\n'
    return codigo_zpl

#Busca os valores na planilha
def busca_linha(inicio, planilha):
    while True:
        # Procura a linha onde a variável "inicio" está
        row_number = 0
        for row in planilha.iter_rows():
            row_number += 1
            if row[0].value == inicio:
                return row_number + 1
        else:
            return 1


#Caminho da Planilha gerada pelo exel
workbook = openpyxl.load_workbook('Q:/#/Suporte-TI/Fred/Automacao-logistica/ENDEREÇOS.xlsx')
sheet = workbook.active


def funcao():
    # Inicia com a célula A2
    row_number = 2
    var = sheet.cell(row=row_number, column=1).value

    # Loop enquanto a célula atual não estiver vazia
    while var is not None:
        concatena = '05' + var
        formatado = "-".join([var[:2], var[2:5], var[5], var[6:]])

        # Layout da etiqueta
        codigo_zpl = codigo_impressora(concatena, formatado)

        impressora(codigo_zpl)

        # Atualiza para a próxima célula da coluna A
        row_number += 1
        var = sheet.cell(row=row_number, column=1).value

    messagebox.showinfo("STATUS", "Etiquetas Geradas")

def abrir_caixa():

    valor = sd.askstring("REIMPRESSÃO", "Digite o endereço:")

    if valor is not None and valor != "":
        existe = busca_linha(valor, sheet)
        if existe != 1:
            concatena = '05' + valor
            formatado = "-".join([valor[:2], valor[2:5], valor[5], valor[6:]])
            codigo_zpl = codigo_impressora(concatena, formatado)
            impressora(codigo_zpl)
            messagebox.showinfo("REIMPRESSÃO", f"O endereço {valor} foi gerado!")
        else:
            resposta = messagebox.askquestion("Endereço não encontrado", "Essa localização não consta na planilha, deseja imprimir mesmo assim?")
            if resposta == "yes":
                concatena = '05' + valor
                formatado = "-".join([valor[:2], valor[2:5], valor[5], valor[6:]])
                codigo_zpl = codigo_impressora(concatena, formatado)
                impressora(codigo_zpl)
                messagebox.showinfo("REIMPRESSÃO", f"O endereço {valor} foi gerado!")
    else:
        messagebox.showwarning("Erro", "Nenhum dado foi inserido.")

def abrir_caixas():
    valores = sd.askstring("Intervalo ", "Digite dois valores separados por vírgula:")
    if valores is not None and valores != "":
        inicio, fim = valores.split(",")

        # Verifica Posição Inicial
        posInicial = busca_linha(inicio, sheet)

        if posInicial != 1:
            # Verifica Posição Final
            porFinal = busca_linha(fim, sheet)

            if porFinal != 1:
                # Inicia com a célula A2
                row_number = posInicial
                var = sheet.cell(row=row_number, column=1).value
                i = posInicial
                # Loop enquanto a célula atual não estiver vazia
                while i <= porFinal and var is not None:
                    concatena = '05' + var
                    formatado = "-".join([var[:2], var[2:5], var[5], var[6:]])

                    # Layout da etiqueta
                    codigo_zpl = codigo_impressora(concatena, formatado)

                    impressora(codigo_zpl)

                    # Atualiza para a próxima célula da coluna A
                    row_number += 1
                    var = sheet.cell(row=row_number, column=1).value
                    i += 1

                # Verifica se o loop parou por causa da célula vazia ou por atingir a posição final
                if i <= porFinal:
                    messagebox.showwarning("Erro", "Por favor, insira uma localização Final Válida!")
                else:
                    messagebox.showinfo("Intervalo", f"Etiquetas de {inicio.strip()} até {fim.strip()} foram geradas!")
            else:
                messagebox.showwarning("Erro", "Por favor, insira uma localização Final Válida!")
        else:
            messagebox.showwarning("Erro", "Por favor, insira uma localização Inicial Válida!")
    else:
        messagebox.showwarning("Erro", "Nenhum dado foi inserido.")



def encerrar():
    root.destroy()

root = tk.Tk()
root.geometry("300x300")  # aumenta a altura da janela para 400 pixels
root.title("Gerador de Etiquetas")


btn1 = tk.Button(root, text="Gerar Todos Endereços", command=funcao, width=20)
btn2 = tk.Button(root, text="Reemprimir", command=abrir_caixa, width=20)
btn3 = tk.Button(root, text="Imprimir Intervalo", command=abrir_caixas, width=20)
btn4 = tk.Button(root, text="Encerrar programa", command=encerrar, width=20)

btn1.pack(side="top", padx=10, pady=10)
btn2.pack(side="top", padx=10, pady=10)
btn3.pack(side="top", padx=10, pady=10)
btn4.pack(side="top", padx=10, pady=10)



# Define o caminho do arquivo de imagem do ícone
caminho_icone = "C:/Users/frederico.almeida/PycharmProjects/Etq_v2/tech.ico"

# Define o ícone da janela do Tkinter
root.iconbitmap(caminho_icone)


# Crie um canvas e adicione-o à sua janela
canvas = Canvas(root, width=300, height=300)
canvas.pack()

# Adicione o texto de copyright no canto inferior direito
copyright = canvas.create_text(150, 100, text="© 2023 Frederico Almeida. Todos os direitos reservados.")

root.mainloop()
